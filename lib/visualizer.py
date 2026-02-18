from rdkit import Chem
from rdkit.Chem.Draw import rdMolDraw2D
import logging
from typing import Dict, List, Set, Optional
import pandas as pd
import os
import sys

# 动态添加 lib 路径以支持导入 (Dynamically add lib path to support imports)
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

try:
    import sugar_utils
except ImportError:
    # Fallback if running from root
    from lib import sugar_utils

# 配置日志 (Configure logging)
logger = logging.getLogger(__name__)

class StructureVisualizer:
    """
    负责糖脂结构的解析与可视化 (Handles structure parsing and visualization of glycolipids)
    核心功能：基于 sugar_utils 识别糖环，区分糖基(Glycan)与苷元(Aglycone)，并进行染色可视化。
    """
    
    def __init__(self):
        pass

    def analyze_glycolipid(self, smiles: str, output_path: str = None) -> bool:
        """
        解析糖脂结构并生成可视化图像 (Analyze glycolipid structure and generate visualization)
        
        Args:
            smiles: 分子的 SMILES 字符串
            output_path: 图片保存路径 (可选)
            
        Returns:
            bool: 是否成功生成 (Success status)
        """
        mol = Chem.MolFromSmiles(smiles)
        if not mol:
            logger.error(f"Invalid SMILES: {smiles}")
            return False

        try:
            # 1. 使用 sugar_utils 识别糖单元 (Identify Sugar Units using sugar_utils)
            sugar_units, atom_to_sugar = sugar_utils.get_sugar_units(mol)
            # linkages = sugar_utils.find_glycosidic_linkages(mol, sugar_units, atom_to_sugar) # Unused in drawing for now
            
            # 2. 收集所有属于“糖部分”的原子 (Collect all atoms belonging to the "Sugar Part")
            all_sugar_unit_atoms: Set[int] = set() # Ring + Exocyclic carbons (e.g. C6)
            
            for unit in sugar_units:
                all_sugar_unit_atoms.update(unit['position_map'].keys())
                
            # 4. Classification (Delegated to sugar_utils)
            classification = sugar_utils.classify_sugar_parts(mol)
            sugar_ring_atoms = classification['sugar_ring_atoms']
            sugar_substituent_atoms = classification['sugar_substituent_atoms']
            aglycone_atoms = classification['aglycone_atoms']
            all_sugar_framework_atoms = classification.get('all_sugar_framework_atoms', set()) # C6 etc.

            # 5. 准备着色方案 (Prepare Coloring Scheme)
            # 糖环 (Sugar Ring): 红色 (Red) (1.0, 0.6, 0.6)
            # 糖基团/连接 (Substituents/Linkages): 黄色 (Yellow) (1.0, 1.0, 0.4)
            # 苷元 (Aglycone): 蓝色 (Blue) (0.6, 0.8, 1.0)
            
            highlight_atoms = list(range(mol.GetNumAtoms()))
            highlight_atom_colors: Dict[int, tuple] = {}
            
            for idx in range(mol.GetNumAtoms()):
                if idx in sugar_ring_atoms:
                    highlight_atom_colors[idx] = (1.0, 0.5, 0.5) # Red
                elif idx in sugar_substituent_atoms:
                    highlight_atom_colors[idx] = (1.0, 1.0, 0.4) # Yellow
                elif idx in all_sugar_framework_atoms:
                     # Exocyclic carbons (C6) -> Yellow to distinguish from ring, or light red?
                     # User wants "Sugar Ring" vs "Substituents".
                     # C6 is part of the sugar framework. Let's make it Light Red/Pink or same as Ring?
                     # Usually C6 is 'part of the sugar'.
                     # Let's keep it consistent: Sugar Framework = Red.
                     # But current request: Substituents = Yellow.
                     # Let's make C6 Red (Sugar) unless it's modified.
                     highlight_atom_colors[idx] = (1.0, 0.5, 0.5) # Red
                elif idx in aglycone_atoms:
                    highlight_atom_colors[idx] = (0.5, 0.8, 1.0) # Blue
                else:
                    # Fallback (e.g. ions, or unaccounted atoms) -> Blue (Aglycone-like)
                    highlight_atom_colors[idx] = (0.5, 0.8, 1.0)

            # 6. 绘制图像 (Draw Image)
            d2d = rdMolDraw2D.MolDraw2DCairo(800, 600) # Increased resolution
            
            # 设置绘图选项 (Drawing Options)
            dopts = d2d.drawOptions()
            dopts.useBWAtomPalette() # Black and White atoms
            dopts.padding = 0.05
            
            d2d.DrawMolecule(mol, highlightAtoms=highlight_atoms, highlightAtomColors=highlight_atom_colors)
            d2d.FinishDrawing()
            
            if output_path:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                d2d.WriteDrawingText(output_path)
            
            return True
            
        except Exception as e:
            logger.error(f"Error analyzing/drawing molecule: {e}", exc_info=True)
            return False

    # Removed internal classify_atoms method as it is now in sugar_utils
    def batch_process_from_file(self, file_path: str, output_dir: str = "images/batch/") -> None:
        """
        批量处理文件中的所有 SMILES，并将生成的图片嵌入到 Excel 中 (Batch process all SMILES and embed images into Excel)
        Structure_Image will be placed in the FIRST column (Column A).
        """
        logger.info(f"Starting batch processing from {file_path}...")
        
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return

        os.makedirs(output_dir, exist_ok=True)
            
        try:
            # 2. 准备 OpenPyXL 进行编辑 (Prepare OpenPyXL for editing)
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(file_path)
            # Process all sheets
            sheets = wb.sheetnames
            
            for sheet_name in sheets:
                ws = wb[sheet_name]
                
                # Check Header
                header_row_vals = [cell.value for cell in ws[1]]
                
                # Find SMILES column (0-based index in this list)
                try:
                    # Case insensitive search
                    smiles_idx_list = [i for i, h in enumerate(header_row_vals) if h and str(h).lower() in ['smiles', 'canonical_smiles']]
                    if not smiles_idx_list:
                        logger.warning(f"No SMILES column in sheet {sheet_name}")
                        continue
                    original_smiles_col_idx = smiles_idx_list[0] + 1 # 1-based for openpyxl
                except ValueError:
                    continue

                # Insert Image Column at A if not present
                current_img_col_idx = 1
                if header_row_vals[0] != 'Structure_Image':
                    logger.info(f"Inserting Structure_Image column at A in {sheet_name}...")
                    ws.insert_cols(1)
                    ws.cell(row=1, column=1, value='Structure_Image')
                    # Shift SMILES column index
                    original_smiles_col_idx += 1
                
                # Iterate rows
                success_count = 0
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    try:
                        smiles_cell = row[original_smiles_col_idx - 1] # tuple index is 0-based
                        smiles = smiles_cell.value
                    except IndexError:
                        continue

                    if not smiles or not isinstance(smiles, str):
                        continue
                    
                    # Generate ID for filename
                    img_name = f"struct_{sheet_name}_{i}"
                    output_image_path = os.path.join(output_dir, f"{img_name}.png")
                    
                    # Generate and Embed
                    if self.analyze_glycolipid(smiles, output_image_path):
                        success_count += 1
                        try:
                            img = XLImage(output_image_path)
                            img.width = 300
                            img.height = 200
                            
                            cell_loc = ws.cell(row=i, column=current_img_col_idx).coordinate
                            ws.add_image(img, cell_loc)
                            ws.row_dimensions[i].height = 150
                        except Exception as e:
                            logger.warning(f"Embedding failed for row {i}: {e}")
                
                # Adjust column width for A
                ws.column_dimensions['A'].width = 40
                logger.info(f"Embedded {success_count} images in {sheet_name}")

            wb.save(file_path)
            logger.info(f"Batch processing complete. Saved to {file_path}")
            
        except Exception as e:
            logger.error(f"Error in batch processing: {e}", exc_info=True)

if __name__ == "__main__":
    # Test with a sample SMILES
    test_smiles = "CO[C@@H]1O[C@H](CO[C@@H]2O[C@H](COC(=O)OCC3c4ccccc4-c4ccccc43)[C@@H](OCc3ccccc3)[C@H](OCc3ccccc3)[C@H]2OC(=O)c2ccccc2)[C@@H](OCc2ccccc2)[C@H](OCc2ccccc2)[C@H]1OC(=O)c1ccccc1"
    viz = StructureVisualizer()
    viz.analyze_glycolipid(test_smiles, "images/test_sugar_utils_viz.png")
    print("Test image generated: images/test_sugar_utils_viz.png")

